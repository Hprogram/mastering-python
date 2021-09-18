import os, sys
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
import librosa
import numpy as np
from FeatureGetter import *
from Util.logger import *
from LearningModule import TuningData
import openpyxl
import random

class ExcelAugmantation:
    def __init__(self, filePath):
        wb = openpyxl.load_workbook(filePath)
        self.data = {}
        self.saveFilePath = filePath.replace(".xlsx", "_Test.xlsx")
        self.sheetList = []
        
        sheetCounter = -1
        for sheetName in wb.get_sheet_names():
            ws = wb.get_sheet_by_name(sheetName)
            self.sheetList.append(ws.title)
            rowNoneCounter = 0
            sheetCounter += 1
            for r in ws.rows:
                fileName = r[0].value
                # if type(fileName) == type(None): 
                try:
                    if fileName.find(".wav") == -1:
                        rowNoneCounter+=1
                        if rowNoneCounter > 4: break
                        continue
                except AttributeError:
                    rowNoneCounter += 1
                    if rowNoneCounter > 4: break
                    continue
                
                rowNoneCounter = 0
                self.data[fileName] = { 'createData' : r[2].value,
                                        'threshold':r[3].value,
                                        'ratio':r[4].value,
                                        'attack':r[5].value,
                                        'release':r[6].value,
                                        'gain':r[7].value,
                                        'eq1':r[8].value,
                                        'eq2':r[12].value,
                                        'eq3':r[16].value,
                                        'sheetCounter':sheetCounter}
                self.data[fileName.replace(".wav", "_v2.wav")] = { 
                                        'createData' : r[2].value,
                                        'threshold':round(random.uniform(r[3].value, r[3].value*1.2), 1) ,
                                        'ratio':round(random.uniform(r[4].value*0.8, r[4].value*1.2), 1),
                                        'attack':round(random.uniform(r[5].value*0.8, r[5].value*1.2), 1),
                                        'release':round(random.uniform(r[6].value*0.8, r[6].value*1.2), 1),
                                        'gain':round(random.uniform(r[7].value*0.8, r[7].value), 1),
                                        'eq1':r[8].value,
                                        'eq2':r[12].value,
                                        'eq3':r[16].value,
                                        'sheetCounter':sheetCounter}                                        
        
    def writeAugAnswer(self):
        wb = openpyxl.Workbook()
        wsHipop = wb.active
        wsHipop.title = self.sheetList[0]
        wsHipop['A1'] = "Hip"
        
        wsRnB = wb.create_sheet()
        wsRnB.title = self.sheetList[1]
        wsRnB['A1'] = "wsRnB"

        wsRock = wb.create_sheet()
        wsRock.title = self.sheetList[2]
        wsRock['A1'] = "wsRock"

        wsPop = wb.create_sheet()
        wsPop.title = self.sheetList[3]
        wsPop['A1'] = "wsPop"

        wsClassic = wb.create_sheet()
        wsClassic.title = self.sheetList[4]
        wsClassic['A1'] = "wsClassic"

        wsJazz = wb.create_sheet()
        wsJazz.title = self.sheetList[5]
        wsJazz['A1'] = "wsJazz"   
        
        sheetCounter = 0
        row_index = 1
        for key, value in self.data.items():
            wsHipop.cell(row=row_index, column=1).value = key
            wsHipop.cell(row=row_index, column=2).value = value['createData']
            wsHipop.cell(row=row_index, column=3).value = value['threshold']
            wsHipop.cell(row=row_index, column=4).value = value['ratio']
            wsHipop.cell(row=row_index, column=5).value = value['attack']
            wsHipop.cell(row=row_index, column=6).value = value['release']
            wsHipop.cell(row=row_index, column=7).value = value['gain']
            wsHipop.cell(row=row_index, column=8).value = value['eq1']
            wsHipop.cell(row=row_index, column=9).value = value['eq2']
            wsHipop.cell(row=row_index, column=10).value = value['eq3']
            row_index += 1
        

        wb.save(filename = self.saveFilePath)
        

if (__name__ == '__main__'):
    excelAug = ExcelAugmantation('./FeatureExtract/LearnData_22k_mono/200328DEEP MASTER_v2.xlsx')
    excelAug.writeAugAnswer()